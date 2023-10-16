from pathlib import Path
from openpyxl import load_workbook
from src.file_format import dump_csv


def excel2csv(excel_file_path, main_sheet=None):
    excel_file_path = Path(excel_file_path)

    table = load_excel(excel_file_path, main_sheet)

    csv_file_path = (
        excel_file_path.parent /
        f'{excel_file_path.stem}.csv')

    dump_csv(csv_file_path, table)


def load_excel(excel_file_path, main_sheet=None):

    workbook = load_workbook(str(excel_file_path), read_only=True)

    sheet = None
    for sheet_name in workbook.sheetnames:
        if main_sheet:
            if sheet_name != main_sheet:
                continue
            sheet = workbook[sheet_name]
            break
        else:
            sheet = workbook[sheet_name]
            break

    sheet = workbook[sheet_name]
    table = []

    header = []
    for idx, i in enumerate(sheet):
        if idx == 0:
            for j in i:
                header.append(j.value)
            continue

        row = []
        for j in i:
            value = j.value
            if j.number_format and j.number_format.endswith('%'):
                value = f"{value * 100}%"
            if value is None:
                value = ''
            row.append(value)

        # Skip blank row
        if not any(row):
            continue
        table.append(dict(zip(header, row)))

    return table
